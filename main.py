from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
import json
from openpyxl import load_workbook
from transformers import RagTokenizer, RagTokenForGeneration
from huggingface_hub import login
import os

app = FastAPI()

frontend_dir = os.path.join(os.path.dirname(__file__), 'frontend')

if not os.path.exists(frontend_dir):
    raise RuntimeError(f"Directory '{frontend_dir}' does not exist")

app.mount("/static", StaticFiles(directory=frontend_dir), name="static")

# Login to Hugging Face API
login("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")

# Load the RAG model and tokenizer
try:
    tokenizer = RagTokenizer.from_pretrained("facebook/rag-token-base")
    model = RagTokenForGeneration.from_pretrained("facebook/rag-token-base")
except Exception as e:
    raise RuntimeError(f"Failed to load RAG model or tokenizer: {str(e)}")

def excel_to_json(file: UploadFile) -> str:
    try:
        workbook = load_workbook(filename=file.file, data_only=True)
        sheet = workbook.active
        data = []
        headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))
        return json.dumps(data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing Excel file: {str(e)}")

def generate_answer(json_data: str, question: str) -> str:
    try:
        context_list = json.loads(json_data)
        
        # Convert JSON data to a context string
        context = " ".join([str(item) for item in context_list])
        
        # Truncate the context to fit within the model's max input length
        context = context[:512]  # Adjust this value if needed

        # Tokenize the input and question
        inputs = tokenizer(question, context, return_tensors="pt", truncation=True)

        # Ensure input_ids and attention_mask are not None
        if inputs.get("input_ids") is None or inputs.get("attention_mask") is None:
            raise ValueError("Tokenization resulted in None for input_ids or attention_mask")

        # Generate the answer
        outputs = model.generate(inputs["input_ids"], attention_mask=inputs["attention_mask"])

        # Decode the generated text
        answer = tokenizer.decode(outputs[0], skip_special_tokens=True)
        return answer
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating answer: {str(e)}")

@app.get("/")
async def serve_index():
    return FileResponse(os.path.join(frontend_dir, "index.html"))

@app.post("/ask_question/")
async def ask_question(question: str = Form(...), file: UploadFile = File(...)):
    try:
        json_data = excel_to_json(file)
        answer = generate_answer(json_data, question)
        return {"answer": answer}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error in ask_question endpoint: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
